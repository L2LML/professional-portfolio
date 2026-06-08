"""
build_appeal_pdf.py — Generate tax appeal exhibit PDF from the tracker spreadsheet.

Reads the "Master Exhibit Log" sheet (formerly "Exhibit Log"), loads images/PDFs
from the Fairlane receipts folder, and builds a PDF where:
  • Each exhibit gets its own page(s) with its title in the header
  • Page numbers are written back into the "PDF Page" column of the tracker
  • Status filter: ✅ SUBMITTED/OBTAINED/ATTACHED/RECEIVED  +  📸/📄 found/new

Usage:
    python3 build_appeal_pdf.py
    python3 build_appeal_pdf.py --cols 2 --output Appeal_Exhibits.pdf --preview
"""

import argparse
import io
import re
import subprocess
import sys
from pathlib import Path

# ── Default paths ─────────────────────────────────────────────────────────────

TRACKER_PATH = Path(
    "/Volumes/Lisa 4TB/A Documents/Tax return/2025/Tax appeal"
    "/files. tax appeal emails/Fairlane receipts"
    "/Lewandowski_Appeal_Tracker_v8.xlsx"
)

PHOTOS_PATH = Path(
    "/Volumes/Lisa 4TB/A Documents/Tax return/2025/Tax appeal"
    "/files. tax appeal emails/Fairlane receipts"
)

OUTPUT_PATH = Path("/Users/lisamlewandowski/Projects/PDF image placement/Appeal_Exhibits.pdf")

# ── Status filter ──────────────────────────────────────────────────────────────

def include_status(status: str) -> bool:
    """Return True for exhibits that belong in the PDF."""
    if not status or status == "nan":
        return False
    return (status.startswith("✅") or status.startswith("🆕") or
            status.startswith("📸") or status.startswith("📄"))

# ── Page / layout constants ────────────────────────────────────────────────────

PAGE_W  = 8.5 * 72
PAGE_H  = 11.0 * 72
MARGIN  = 36
HEADER_H = 34
FOOTER_H = 20
GAP      = 10
CAPTION_H = 26

IMG_EXTS   = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}
HEIC_EXTS  = {".heic", ".heif"}
PDF_EXTS   = {".pdf"}

# ── File index ─────────────────────────────────────────────────────────────────

def build_file_index(photos_root: Path) -> dict:
    """Map lowercase filename stem → full Path for every file in the tree."""
    index = {}
    for p in photos_root.rglob("*"):
        if p.is_file() and not p.name.startswith("."):
            index[p.stem.lower()] = p
            index[p.name.lower()] = p
    return index


def resolve_files(image_files_cell: str, photos_root: Path, index: dict) -> list:
    """
    Resolve paths from the File Path / Notes cell into a list of Paths.
    Handles: semicolons, newlines, absolute paths, relative paths, stem-only fallback.
    """
    if not image_files_cell or str(image_files_cell).strip() in ("", "nan"):
        return []

    # Split on both semicolons and newlines (tracker uses \n between multiple paths)
    raw_paths = re.split(r'[;\n]+', str(image_files_cell))

    paths = []
    for raw in raw_paths:
        raw = raw.strip()
        if not raw:
            continue

        # 1. Absolute path — use directly if it exists
        p = Path(raw)
        if p.is_absolute() and p.exists():
            paths.append(p)
            continue

        # 2. Relative path inside photos_root
        candidate = photos_root / raw
        if candidate.exists():
            paths.append(candidate)
            continue

        # 3. Stem / filename lookup in full index
        stem = p.stem.lower()
        name = p.name.lower()
        if name in index:
            paths.append(index[name])
        elif stem in index:
            paths.append(index[stem])
        else:
            print(f"  [warning] File not found: {raw}")

    return paths

# ── Image loaders ──────────────────────────────────────────────────────────────

def load_image_bytes(path: Path) -> list:
    """
    Return a list of (caption, jpeg_bytes) for a given file.
    Images → 1 entry. PDFs → one entry per page.
    """
    from PIL import Image as PILImage

    ext = path.suffix.lower()

    if ext in HEIC_EXTS:
        import pillow_heif
        pillow_heif.register_heif_opener()
        img = PILImage.open(path).convert("RGB")
        return [(path.stem, _pil_to_bytes(img))]

    if ext in IMG_EXTS:
        img = PILImage.open(path).convert("RGB")
        return [(path.stem, _pil_to_bytes(img))]

    if ext in PDF_EXTS:
        return _pdf_to_images(path)

    print(f"  [skip] Unsupported format: {path.name}")
    return []


def _pil_to_bytes(img, quality: int = 72) -> bytes:
    buf = io.BytesIO()
    img.convert("RGB").save(buf, format="JPEG", quality=quality, optimize=True)
    return buf.getvalue()


def _pdf_to_images(path: Path) -> list:
    """Render each PDF page to a JPEG using pypdfium2."""
    try:
        import pypdfium2 as pdfium
    except ImportError:
        sys.exit("Missing pypdfium2. Run:  pip3 install pypdfium2")

    from PIL import Image as PILImage

    doc     = pdfium.PdfDocument(str(path))
    results = []
    for i, page in enumerate(doc):
        bitmap = page.render(scale=1.5)   # 108 dpi — good quality, smaller file
        pil    = bitmap.to_pil()
        caption = f"{path.stem}  (p.{i + 1})" if len(doc) > 1 else path.stem
        results.append((caption, _pil_to_bytes(pil)))
    doc.close()
    return results

# ── PDF builder ────────────────────────────────────────────────────────────────

def build_pdf(sections: list, output: Path, cols: int = 2):
    """
    sections = [
        {
          'title':    'Exhibit D2 — Fireplace / Tree Impact',
          'entries':  [(caption, img_bytes), ...],
          'page_num': set by this function, returned
        },
        ...
    ]
    Returns sections list with 'start_page' filled in (1-based).
    """
    from reportlab.pdfgen import canvas
    from reportlab.lib import colors
    from reportlab.lib.utils import ImageReader

    usable_w    = PAGE_W - 2 * MARGIN
    header_used = HEADER_H + GAP
    footer_used = FOOTER_H + GAP
    usable_img_h = PAGE_H - 2 * MARGIN - header_used - footer_used

    cell_w = (usable_w - (cols - 1) * GAP) / cols
    img_h  = cell_w * 0.75
    cell_h = img_h + CAPTION_H + GAP
    rows_per_page  = max(1, int(usable_img_h / cell_h))
    cells_per_page = rows_per_page * cols
    grid_top       = PAGE_H - MARGIN - header_used

    c = canvas.Canvas(str(output), pagesize=(PAGE_W, PAGE_H))
    c.setTitle("2026 Tax Assessment Appeal — Exhibit Package")
    c.setAuthor("Lisa Marie Lewandowski")

    global_page = 0   # 1-based page counter

    def draw_chrome(title: str):
        """Draw header and footer on the current canvas page."""
        nonlocal global_page
        global_page += 1

        # Header
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(colors.HexColor("#1a1a2e"))
        c.drawString(MARGIN, PAGE_H - MARGIN - 14, title)
        c.setStrokeColor(colors.HexColor("#888888"))
        c.setLineWidth(0.5)
        rule_y = PAGE_H - MARGIN - HEADER_H
        c.line(MARGIN, rule_y, PAGE_W - MARGIN, rule_y)

        # Footer
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.HexColor("#555555"))
        c.line(MARGIN, MARGIN + FOOTER_H, PAGE_W - MARGIN, MARGIN + FOOTER_H)
        c.drawString(MARGIN, MARGIN + 5,
                     "14707 Fairlane St, Livonia MI 48154  |  Parcel 46 082 06 0049 000  |  Lisa M. Lewandowski")
        c.drawRightString(PAGE_W - MARGIN, MARGIN + 5, f"Page {global_page}")

    for section in sections:
        title   = section["title"]
        entries = section["entries"]

        if not entries:
            section["start_page"] = None
            continue

        section["start_page"] = global_page + 1
        cell_idx = 0

        draw_chrome(title)

        for caption, img_bytes in entries:
            if cell_idx > 0 and cell_idx % cells_per_page == 0:
                c.showPage()
                draw_chrome(title)

            row = (cell_idx % cells_per_page) // cols
            col = (cell_idx % cells_per_page) % cols

            x = MARGIN + col * (cell_w + GAP)
            y = grid_top - (row + 1) * img_h - row * (CAPTION_H + GAP)

            c.setFillColor(colors.HexColor("#f2f2f2"))
            c.rect(x, y, cell_w, img_h, fill=1, stroke=0)

            try:
                ir      = ImageReader(io.BytesIO(img_bytes))
                iw, ih  = ir.getSize()
                scale   = min(cell_w / iw, img_h / ih)
                dw, dh  = iw * scale, ih * scale
                c.drawImage(ir,
                            x + (cell_w - dw) / 2,
                            y + (img_h  - dh) / 2,
                            width=dw, height=dh,
                            preserveAspectRatio=True, mask="auto")
            except Exception as e:
                print(f"  [warning] Could not render '{caption}': {e}")

            c.setFont("Helvetica", 8)
            c.setFillColor(colors.HexColor("#333333"))
            max_chars = int(cell_w / (8 * 0.52))
            short = (caption[:max_chars - 1] + "…") if len(caption) > max_chars else caption
            c.drawCentredString(x + cell_w / 2, y - CAPTION_H + 8, short)

            cell_idx += 1

        c.showPage()

    c.save()
    return sections

# ── Tracker read / write ───────────────────────────────────────────────────────

def read_tracker(tracker_path: Path) -> list:
    """Read Master Exhibit Log and return list of exhibit dicts."""
    import pandas as pd
    import openpyxl

    wb = openpyxl.load_workbook(tracker_path, read_only=True, data_only=True)

    # Support both old sheet name ("Exhibit Log") and current ("Master Exhibit Log")
    sheet_name = "Master Exhibit Log" if "Master Exhibit Log" in wb.sheetnames else "Exhibit Log"
    ws = wb[sheet_name]

    header_excel_row = None
    col_image_idx    = None
    col_page_idx     = None

    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[0].value == "Exhibit":
            header_excel_row = row[0].row
            for cell in row:
                # Support both old ("Image Files") and current ("File Path / Notes") column name
                if cell.value in ("Image Files", "File Path / Notes"):
                    col_image_idx = cell.column - 1
                if cell.value == "PDF Page":
                    col_page_idx  = cell.column - 1
            break
    wb.close()

    if header_excel_row is None:
        raise ValueError(f"Could not find header row in sheet '{sheet_name}'.")

    header_pandas = header_excel_row - 1

    df = pd.read_excel(tracker_path, sheet_name=sheet_name,
                       header=header_pandas, dtype=str)

    # Map columns by NAME, not position — tracker column order may vary
    col_map = {v: v for v in df.columns}
    rename = {}

    # Find each expected column by partial name match
    for col in df.columns:
        cl = str(col).lower()
        if cl == 'exhibit':
            rename[col] = 'exhibit'
        elif 'document' in cl or 'description' in cl or 'source' in cl:
            rename[col] = 'source'
        elif 'category' in cl:
            rename[col] = 'category'
        elif 'amount' in cl:
            rename[col] = 'amount'
        elif 'status' in cl:
            rename[col] = 'status'
        elif 'date' in cl:
            rename[col] = 'date'
        elif 'file path' in cl or 'image files' in cl or 'notes' in cl:
            rename[col] = 'image_files'
        elif 'impact' in cl:
            rename[col] = 'impact'
        elif 'pdf page' in cl:
            rename[col] = 'pdf_page'

    df = df.rename(columns=rename)

    exhibits = []
    for _, row in df.iterrows():
        ex = str(row.get("exhibit", "")).strip()
        if not ex or ex == "nan" or len(ex) > 5:
            continue
        exhibits.append({
            "exhibit":     ex,
            "source":      str(row.get("source", "")).strip(),
            "category":    str(row.get("category", "")).strip(),
            "status":      str(row.get("status", "")).strip(),
            "image_files": str(row.get("image_files", "")).strip()
                           if "image_files" in row.index else "",
        })
    return exhibits


def write_page_numbers(tracker_path: Path, page_map: dict):
    """Write start_page values back into the PDF Page column (adds column if missing)."""
    import openpyxl
    wb = openpyxl.load_workbook(tracker_path)

    sheet_name = "Master Exhibit Log" if "Master Exhibit Log" in wb.sheetnames else "Exhibit Log"
    ws = wb[sheet_name]

    # Find header row and PDF Page column; add it if missing
    col_page   = None
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row[0].value == "Exhibit":
            header_row = row[0].row
            for cell in row:
                if cell.value == "PDF Page":
                    col_page = cell.column
            break

    if col_page is None and header_row:
        # Add "PDF Page" as the next empty column
        col_page = ws.max_column + 1
        ws.cell(row=header_row, column=col_page, value="PDF Page")
        print("  Added 'PDF Page' column to tracker.")

    if col_page is None:
        print("  [warning] Could not add PDF Page column — skipping write-back.")
        wb.close()
        return

    for row in ws.iter_rows(min_row=header_row + 1 if header_row else 5):
        ex = str(row[0].value).strip() if row[0].value else ""
        if ex in page_map and page_map[ex] is not None:
            ws.cell(row=row[0].row, column=col_page, value=page_map[ex])

    wb.save(tracker_path)
    print("  PDF Page numbers written back to tracker.")

# ── Main ───────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Build appeal exhibit PDF from tracker.")
    p.add_argument("--tracker", type=Path, default=TRACKER_PATH)
    p.add_argument("--photos",  type=Path, default=PHOTOS_PATH)
    p.add_argument("--output",  type=Path, default=OUTPUT_PATH)
    p.add_argument("--cols",    type=int,  default=2,
                   help="Images per row (default: 2)")
    p.add_argument("--preview", action="store_true",
                   help="Open PDF automatically after saving")
    return p.parse_args()


def main():
    args = parse_args()

    if not args.tracker.exists():
        sys.exit(f"Tracker not found: {args.tracker}")
    if not args.photos.exists():
        sys.exit(f"Photos folder not found: {args.photos}")

    print(f"Reading tracker: {args.tracker.name} …")
    exhibits = read_tracker(args.tracker)
    print(f"  {len(exhibits)} exhibits found.")

    print(f"Indexing files in {args.photos.name} …")
    file_index = build_file_index(args.photos)
    print(f"  {len(file_index)//2} files indexed.")

    sections = []
    skipped  = []
    for ex in exhibits:
        if not include_status(ex["status"]):
            skipped.append(ex["exhibit"])
            continue

        title   = f"Exhibit {ex['exhibit']} — {ex['category']}"
        files   = resolve_files(ex["image_files"], args.photos, file_index)

        entries = []
        for fp in files:
            loaded = load_image_bytes(fp)
            entries.extend(loaded)
            print(f"  Exhibit {ex['exhibit']:4}  {fp.name[:55]:<55}  {len(loaded)} image(s)")

        if not entries:
            print(f"  Exhibit {ex['exhibit']:4}  [no images — skipping]")

        if entries:
            sections.append({
                "exhibit": ex["exhibit"],
                "title":   title,
                "entries": entries,
            })

    if skipped:
        print(f"\nSkipped (status not included): {', '.join(skipped)}")

    total_images = sum(len(s["entries"]) for s in sections)
    if total_images == 0:
        sys.exit("\nNo images found. Check file paths in the 'File Path / Notes' column.")

    print(f"\nBuilding PDF ({args.cols} col, {total_images} images, {len(sections)} exhibits) …")
    sections = build_pdf(sections, args.output, cols=args.cols)

    page_map = {s["exhibit"]: s.get("start_page") for s in sections}
    write_page_numbers(args.tracker, page_map)

    print(f"\n✅  Saved → {args.output}")
    print(f"    PDF page numbers written back to tracker.")

    if args.preview:
        import platform
        opener = {"darwin": "open"}.get(platform.system().lower(), "xdg-open")
        subprocess.Popen([opener, str(args.output)])


if __name__ == "__main__":
    main()
